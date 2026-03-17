"""
train.py  -  Standalone fine-tuner for university corpus QA
============================================================
No package needed. Just run:

    python train.py

Tuned for RTX 3050 6GB (Ampere, CUDA).
Saves ONE clean final model to models/club-slm/ -- ready for ask.py.

Edit the CONFIG block below if your paths differ.
"""
from __future__ import annotations

import os, sys
# Force UTF-8 stdout/stderr on Windows (fixes cp1252 UnicodeEncodeError)
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import csv
import inspect
import json
import math
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass
from pathlib import Path

# -----------------------------------------------------------------------------
# CONFIG  <-- edit these if your paths differ
# -----------------------------------------------------------------------------
CORPUS_PATH  = Path(r"J:\MAJOR PROJECT 2026\MAIN\data\clean_corpus.txt")
DATASET_PATH = Path("data/python/train_qa.jsonl")
OUTPUT_DIR   = Path("models/club-slm")          # ask.py reads from here
BASE_MODEL   = "Qwen/Qwen2.5-0.5B-Instruct"     # ~1GB download, cached after first run

# -- RTX 3050 6GB optimised settings -----------------------------------------
# Ampere GPU: bf16 used automatically (faster + more stable than fp16)
# Effective batch = BATCH_SIZE x GRAD_ACCUM = 2 x 8 = 16
EPOCHS       = 5        # more epochs = better recall on your corpus
BATCH_SIZE   = 2        # safe for 6GB VRAM with Qwen 0.5B + bf16
GRAD_ACCUM   = 8        # keeps effective batch = 16
LR           = 3e-5     # slightly higher LR for small model + small dataset
MAX_LENGTH   = 256      # covers all your QA pairs; raise to 384 if answers get cut
WARMUP_RATIO = 0.05
# -----------------------------------------------------------------------------

try:
    import torch
    from datasets import Dataset
    from transformers import (
        AutoModelForCausalLM,
        AutoTokenizer,
        DataCollatorForLanguageModeling,
        Trainer,
        TrainingArguments,
    )
except ImportError as e:
    sys.exit(
        f"[ERROR] Missing library: {e}\n"
        "Run:  pip install torch transformers datasets scikit-learn accelerate"
    )

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed -- metrics will be saved as .txt only.")
    print("[WARN] To get Excel output:  pip install openpyxl")


# -----------------------------------------------------------------------------
# Text utilities  (inlined -- no shared.py needed)
# -----------------------------------------------------------------------------

@dataclass
class Chunk:
    chunk_id: int
    text: str
    headings: list


def normalize_text(value: str) -> str:
    value = value.replace("\r", "").replace("\t", " ")
    # Fix common mojibake sequences
    value = value.replace("\xe2\x80\x99", "'").replace("\xe2\x80\x9c", '"')
    value = value.replace("\xe2\x80\x9d", '"').replace("\xe2\x80\x93", "-")
    value = value.replace("\xe2\x80\x94", "-")
    return value.strip()


def is_heading_line(line: str) -> bool:
    return len(line.split()) <= 18 and bool(re.fullmatch(r"[A-Z0-9 .:&'()/,-]+", line))


def split_into_chunks(text: str, max_words: int = 110) -> list:
    chunks: list = []
    current_lines: list = []
    current_headings: list = []
    current_words = 0

    def flush():
        nonlocal current_lines, current_words
        if not current_lines:
            return
        chunks.append(Chunk(
            chunk_id=len(chunks) + 1,
            text=" ".join(current_lines).strip(),
            headings=list(dict.fromkeys(current_headings)),
        ))
        current_lines.clear()
        current_words = 0

    for raw in normalize_text(text).split("\n"):
        line = raw.strip()
        if not line:
            continue
        if re.fullmatch(r"--- PAGE \d+ ---", line, flags=re.IGNORECASE):
            flush()
            current_headings = [h for h in current_headings if not h.startswith("--- PAGE")]
            current_headings.append(line)
            continue
        if line.startswith("DOCUMENT:") or re.match(r"=+\s*DOCUMENT:", line):
            flush()
            current_headings = [re.sub(r"=+", "", line).strip()]
            continue
        if is_heading_line(line):
            if current_words >= max_words // 2:
                flush()
            kept = [h for h in current_headings
                    if h.startswith("DOCUMENT:") or h.startswith("--- PAGE")]
            current_headings = kept + [line]
            continue
        line_words = len(line.split())
        force_break = (bool(re.match(r"^\d+[.)]\s+", line))
                       or bool(re.match(r"^\d+\s+[A-Z]", line)))
        if current_lines and (current_words + line_words > max_words or force_break):
            flush()
        current_lines.append(line)
        current_words += line_words

    flush()
    return chunks


def compact_answer(text: str, max_sentences: int = 2) -> str:
    parts = re.split(r"(?<=[.!?])\s+", normalize_text(text))
    parts = [p.strip() for p in parts if p.strip()]
    return " ".join(parts[:max_sentences]) if parts else normalize_text(text)


# -----------------------------------------------------------------------------
# QA dataset builder  (inlined -- no prepare_dataset.py needed)
# -----------------------------------------------------------------------------

def build_examples(corpus_path: Path) -> list:
    print(f"[DATA] Reading corpus: {corpus_path}", flush=True)
    text = corpus_path.read_text(encoding="utf-8", errors="ignore")
    chunks = split_into_chunks(text)
    print(f"[DATA] {len(chunks)} chunks parsed.", flush=True)

    rows: list = []
    seen: set = set()

    def add(question: str, answer: str, evidence: str, source: str) -> None:
        q, a = question.strip(), answer.strip()
        if not q or not a:
            return
        key = (q.lower(), a.lower())
        if key in seen:
            return
        seen.add(key)
        rows.append({"question": q, "answer": a, "evidence": evidence.strip(), "source": source})

    for chunk in chunks:
        source   = " > ".join(chunk.headings) if chunk.headings else f"chunk-{chunk.chunk_id}"
        evidence = chunk.text

        # -- Field: value pairs (Chairperson: John, Fee: Rs.50000, etc.) ----
        for m in re.finditer(
            r"(?:^|\s)(?:\d+[.)]\s*)?([A-Za-z][A-Za-z &/()'-]{2,40}):\s*([^.;\n][^\n]{1,140})",
            chunk.text,
        ):
            field = re.sub(r"\s+", " ", m.group(1)).strip()
            value = re.sub(r"\s+", " ", m.group(2)).strip(" .")
            if len(field.split()) <= 8 and len(value) >= 2:
                add(f"What is the {field.lower()}?",          f"{field}: {value}", evidence, source)
                add(f"Tell me about the {field.lower()}.",    f"{field}: {value}", evidence, source)
                if any(w in field.lower() for w in ["chair", "secretary", "treasurer", "mentor", "vice"]):
                    add(f"Who can be the {field.lower()}?",              f"{field}: {value}", evidence, source)
                    add(f"What is the eligibility for {field.lower()}?", f"{field}: {value}", evidence, source)

        # -- Fee / finance patterns ------------------------------------------
        for m in re.finditer(
            r"((?:tuition|hostel|mess|admission|enrollment|enrolment|lab)\s+fee[^.]{0,200}\.)",
            chunk.text, re.IGNORECASE
        ):
            sentence = re.sub(r"\s+", " ", m.group(1)).strip()
            add("What is the fee structure?",        sentence, evidence, source)
            add("What are the fees?",                sentence, evidence, source)
            add("How much is the tuition fee?",      sentence, evidence, source)
            add("What is the hostel fee?",           sentence, evidence, source)

        # -- Rs. / rupee amounts ---------------------------------------------
        for m in re.finditer(
            r"((?:Rs\.?|INR|rupees?)\s*[\d,]+[^.]{0,150}\.)",
            chunk.text, re.IGNORECASE
        ):
            sentence = re.sub(r"\s+", " ", m.group(1)).strip()
            add("What are the financial details?", sentence, evidence, source)

        # -- Heading-based QA ------------------------------------------------
        if chunk.headings:
            topic  = chunk.headings[-1]
            answer = compact_answer(chunk.text, max_sentences=2)
            add(f"What does the document say about {topic.lower()}?", answer, evidence, source)
            add(f"Explain {topic.lower()}.",                           answer, evidence, source)

            topic_l = topic.lower()
            if "responsibilit" in topic_l:
                add("What are the responsibilities of clubs?",                    answer, evidence, source)
                add("What are the responsibilities mentioned in the document?",   answer, evidence, source)
            if "objective" in topic_l or "introduction" in topic_l:
                add("What is the objective of this document?",                    answer, evidence, source)
            if "recruitment" in topic_l:
                add("What is the recruitment process?",                           answer, evidence, source)
            if "placement" in topic_l:
                add("What is the placement process?",                             answer, evidence, source)
                add("What are the placement details?",                            answer, evidence, source)
            if "scholarship" in topic_l or "waiver" in topic_l:
                add("What scholarships are available?",                           answer, evidence, source)
                add("What is the fee waiver policy?",                             answer, evidence, source)
            if "hostel" in topic_l or "accommodat" in topic_l:
                add("What are the hostel facilities?",                            answer, evidence, source)
                add("What are the hostel rules?",                                 answer, evidence, source)
            if "exam" in topic_l or "assessment" in topic_l:
                add("How are exams conducted?",                                   answer, evidence, source)
                add("What is the examination policy?",                            answer, evidence, source)
            if "attendance" in topic_l:
                add("What is the attendance requirement?",                        answer, evidence, source)
                add("What is the minimum attendance?",                            answer, evidence, source)
            if "club" in topic_l or "societ" in topic_l:
                add("What clubs are available?",                                  answer, evidence, source)
                add("How to join a club?",                                        answer, evidence, source)

        # -- Numbered list items ---------------------------------------------
        numbered_items = re.findall(r"(?:^|\s)(\d+[.)]\s+[^\n]{10,140})", chunk.text)
        if numbered_items and chunk.headings:
            answer = " ".join(i.strip() for i in numbered_items[:4])
            add(f"List key points from {chunk.headings[-1].lower()}.", answer, evidence, source)

        # -- Eligibility / criteria sentences -------------------------------
        for m in re.finditer(
            r"((?:student|candidate)s?\s+(?:must|should|need to|are required to)[^.]{0,200}\.)",
            chunk.text, re.IGNORECASE
        ):
            sentence = re.sub(r"\s+", " ", m.group(1)).strip()
            add("What are the eligibility criteria?",        sentence, evidence, source)
            add("What are the requirements for students?",   sentence, evidence, source)

        # -- Policy sentences ------------------------------------------------
        for m in re.finditer(
            r"((?:anti-ragging|ragging|discipline|misconduct|code of conduct)[^.]{0,250}\.)",
            chunk.text, re.IGNORECASE
        ):
            sentence = re.sub(r"\s+", " ", m.group(1)).strip()
            add("What is the anti-ragging policy?",    sentence, evidence, source)
            add("What are the disciplinary rules?",    sentence, evidence, source)

    print(f"[DATA] {len(rows)} QA pairs generated.", flush=True)
    return rows


def load_or_build_dataset(dataset_path: Path, corpus_path: Path) -> list:
    if dataset_path.exists():
        rows = []
        with dataset_path.open("r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
        if rows:
            print(f"[DATA] Loaded {len(rows)} existing QA pairs from {dataset_path}", flush=True)
            return rows

    rows = build_examples(corpus_path)
    dataset_path.parent.mkdir(parents=True, exist_ok=True)
    with dataset_path.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=True) + "\n")
    print(f"[DATA] Saved dataset to {dataset_path}", flush=True)
    return rows


def format_example(row: dict) -> str:
    """Instruction-tuning prompt format -- matches what ask.py expects."""
    return (
        "You are a small domain model trained on one university corpus.\n"
        "Answer briefly and exactly from the corpus.\n\n"
        f"Question: {row['question']}\n"
        f"Answer: {row['answer']}"
    )


# -----------------------------------------------------------------------------
# Device + precision detection
# -----------------------------------------------------------------------------

def detect_device() -> str:
    if torch.cuda.is_available():
        gpu   = torch.cuda.get_device_name(0)
        vram  = torch.cuda.get_device_properties(0).total_memory / 1e9
        cap   = torch.cuda.get_device_capability(0)
        print(f"[GPU] {gpu}  |  VRAM: {vram:.1f} GB  |  Compute: sm_{cap[0]}{cap[1]}", flush=True)
        return "cuda"
    # -- CUDA not detected -- common causes on Windows ------------------------
    # 1. PyTorch CPU-only build installed.  Fix:
    #    pip uninstall torch torchvision torchaudio -y
    #    pip install torch torchvision torchaudio --index-url https://download.pytorch.org/whl/cu121
    # 2. CUDA toolkit not on PATH.  Fix: install CUDA 12.x from nvidia.com
    # 3. Driver too old.  Fix: update NVIDIA driver to 525+
    if torch.backends.mps.is_available():
        print("[GPU] Apple MPS backend", flush=True)
        return "mps"
    print("[WARN] No GPU found -- falling back to CPU (will be slow!)", flush=True)
    print("[WARN] If you have an NVIDIA GPU, install the CUDA build of PyTorch:", flush=True)
    print("[WARN]   pip install torch --index-url https://download.pytorch.org/whl/cu121", flush=True)
    return "cpu"


def pick_precision(device: str) -> tuple:
    """Returns (use_fp16, use_bf16, torch_dtype).
    RTX 3050 is Ampere (sm_86) so bf16 + tf32 are both supported.
    bf16 halves VRAM vs fp32 and is faster than fp16 on Ampere.
    """
    if device != "cuda":
        return False, False, torch.float32
    cap = torch.cuda.get_device_capability(0)
    if cap[0] >= 8:   # Ampere (sm_80+): RTX 30xx, 40xx, A-series
        print("[AMP] bf16 mixed precision  (Ampere -- RTX 3050 detected)", flush=True)
        return False, True, torch.bfloat16
    print("[AMP] fp16 mixed precision  (pre-Ampere GPU)", flush=True)
    return True, False, torch.float16


# -----------------------------------------------------------------------------
# Post-training cleanup: keep ONE clean final model, delete all checkpoints
# -----------------------------------------------------------------------------

def purge_checkpoints(output_dir: Path) -> None:
    """Delete all checkpoint-XXXX folders, keep only the final model files."""
    removed = 0
    for child in output_dir.iterdir():
        if child.is_dir() and re.match(r"checkpoint-\d+", child.name):
            shutil.rmtree(child)
            removed += 1
    if removed:
        print(f"[SAVE] Removed {removed} checkpoint folder(s) -- only final model kept.", flush=True)
    else:
        print("[SAVE] No checkpoint folders to remove.", flush=True)



# -----------------------------------------------------------------------------
# Metrics evaluation + Excel / TXT export
# -----------------------------------------------------------------------------

def _normalise(text: str) -> str:
    """Lowercase, collapse whitespace, strip punctuation edges -- for fair comparison."""
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"^[^a-z0-9]+|[^a-z0-9]+$", "", text)
    return text


def _answer_token_ids(input_ids: list, labels: list) -> tuple[list, list]:
    """
    Return only the tokens that belong to the ANSWER portion.
    During training, question/prompt tokens have label=-100 (ignored).
    We find where labels stop being -100 -- that is the answer start.
    """
    answer_input, answer_label = [], []
    in_answer = False
    for inp, lbl in zip(input_ids, labels):
        if lbl != -100:
            in_answer = True
        if in_answer:
            answer_input.append(inp)
            answer_label.append(lbl)
    return answer_input, answer_label


def compute_eval_metrics(trainer: "Trainer", tok_eval, tokenizer, device: str) -> dict:
    """
    Correct metrics that only measure the ANSWER portion of each sequence:

      Perplexity    -- from HF evaluate() on full sequence (standard practice)
      Token Acc     -- predicted token == label token, answer tokens ONLY
                       (ignores prompt/question tokens which always had label=-100)
      Exact Match   -- normalised generated answer == normalised reference answer
                       (normalised = lowercase, stripped punctuation/whitespace)
      Fuzzy Match   -- answer contains all key words from reference (partial credit)
      Wrong tokens  -- answer tokens the model got wrong
    """
    print("\n[METRICS] Evaluating -- answer-only token accuracy ...", flush=True)

    # -- Loss + perplexity (standard HF) --------------------------------------
    eval_result = trainer.evaluate()
    eval_loss   = eval_result.get("eval_loss", float("nan"))
    perplexity  = math.exp(eval_loss) if eval_loss < 20 else float("inf")

    model = trainer.model
    model.eval()

    total_tokens   = 0
    correct_tokens = 0
    wrong_tokens   = 0
    exact_matches  = 0
    fuzzy_matches  = 0
    total_samples  = 0
    sample_log     = []   # first 10 samples for the report

    from torch.utils.data import DataLoader
    loader = DataLoader(tok_eval, batch_size=1, shuffle=False)

    with torch.inference_mode():
        for batch in loader:
            raw_input_ids = batch["input_ids"][0].tolist()       # (T,)
            raw_labels    = batch["labels"][0].tolist()          # (T,)
            attn          = batch.get("attention_mask")

            # -- Token accuracy: answer portion only -----------------------
            ans_inp, ans_lbl = _answer_token_ids(raw_input_ids, raw_labels)
            if not ans_lbl:
                continue

            ans_inp_t = torch.tensor([ans_inp], dtype=torch.long).to(device)
            ans_attn  = torch.ones_like(ans_inp_t)

            out    = model(input_ids=ans_inp_t, attention_mask=ans_attn)
            logits = out.logits[0]          # (T, V)
            preds  = logits.argmax(dim=-1)  # (T,)

            # shift: position i predicts token i+1
            lbl_t = torch.tensor(ans_lbl, dtype=torch.long).to(device)
            shift_preds  = preds[:-1]
            shift_labels = lbl_t[1:]
            valid_mask   = shift_labels != -100

            n_valid   = valid_mask.sum().item()
            n_correct = ((shift_preds == shift_labels) & valid_mask).sum().item()
            n_wrong   = n_valid - n_correct

            total_tokens   += n_valid
            correct_tokens += n_correct
            wrong_tokens   += n_wrong

            # -- Exact + fuzzy match ---------------------------------------
            # Reference = decode the answer label tokens
            ref_ids  = [t for t in ans_lbl if t != -100]
            ref_str  = tokenizer.decode(ref_ids, skip_special_tokens=True)

            # Prediction = decode what the model actually predicted for those positions
            pred_ids = shift_preds[valid_mask].cpu().tolist()
            pred_str = tokenizer.decode(pred_ids, skip_special_tokens=True)

            ref_norm  = _normalise(ref_str)
            pred_norm = _normalise(pred_str)

            is_exact = (pred_norm == ref_norm)
            if is_exact:
                exact_matches += 1

            # Fuzzy: all words in reference appear somewhere in prediction
            ref_words  = set(ref_norm.split())
            pred_words = set(pred_norm.split())
            is_fuzzy   = len(ref_words) > 0 and ref_words.issubset(pred_words)
            if is_fuzzy:
                fuzzy_matches += 1

            total_samples += 1

            # Log first 10 for the report
            if len(sample_log) < 10:
                sample_log.append({
                    "sample":     total_samples,
                    "reference":  ref_str[:120],
                    "predicted":  pred_str[:120],
                    "exact":      is_exact,
                    "fuzzy":      is_fuzzy,
                    "tok_acc":    f"{n_correct}/{n_valid}",
                })

    token_accuracy = correct_tokens / total_tokens  if total_tokens  else 0.0
    exact_match    = exact_matches   / total_samples if total_samples else 0.0
    fuzzy_match    = fuzzy_matches   / total_samples if total_samples else 0.0

    # -- Per-epoch loss curve --------------------------------------------------
    log_history  = trainer.state.log_history
    epoch_losses, eval_losses = [], []
    for entry in log_history:
        if "loss" in entry and "epoch" in entry:
            epoch_losses.append({
                "epoch":      round(entry["epoch"], 2),
                "train_loss": round(entry["loss"], 6),
                "step":       entry.get("step", ""),
            })
        if "eval_loss" in entry and "epoch" in entry:
            eval_losses.append({
                "epoch":     round(entry["epoch"], 2),
                "eval_loss": round(entry["eval_loss"], 6),
            })

    return {
        "eval_loss":       round(eval_loss,       6),
        "perplexity":      round(perplexity,       4),
        "token_accuracy":  round(token_accuracy,   6),
        "exact_match":     round(exact_match,      6),
        "fuzzy_match":     round(fuzzy_match,      6),
        "total_tokens":    total_tokens,
        "correct_tokens":  correct_tokens,
        "wrong_tokens":    wrong_tokens,
        "total_samples":   total_samples,
        "exact_matches":   exact_matches,
        "fuzzy_matches":   fuzzy_matches,
        "epoch_losses":    epoch_losses,
        "eval_losses":     eval_losses,
        "sample_log":      sample_log,
    }


def save_metrics_txt(path: Path, metrics: dict, meta: dict) -> None:
    """Save a human-readable plain-text metrics report."""
    ppl_note = "GREAT" if metrics["perplexity"] < 5 else "GOOD" if metrics["perplexity"] < 15 else "UNDERFIT"
    tok_note  = "GREAT" if metrics["token_accuracy"] > 0.85 else "GOOD" if metrics["token_accuracy"] > 0.60 else "NEEDS MORE TRAINING"
    ex_note   = "GREAT" if metrics["exact_match"]    > 0.50 else "GOOD" if metrics["exact_match"]    > 0.20 else "LOW (normal for generation)"

    lines = [
        "=" * 66,
        "  TRAINING METRICS REPORT",
        "=" * 66,
        "",
        "-- Model --------------------------------------------------------",
        f"  Base model      : {meta['base_model']}",
        f"  Output dir      : {meta['output_dir']}",
        f"  Device          : {meta['device']}",
        f"  Precision       : {meta['precision']}",
        f"  Train time      : {meta['train_time_min']} min",
        "",
        "-- Dataset ------------------------------------------------------",
        f"  QA pairs total  : {meta['qa_pairs']}",
        f"  Training rows   : {meta['train_rows']}",
        f"  Eval rows       : {meta['eval_rows']}",
        f"  Epochs          : {meta['epochs']}",
        f"  Batch size      : {meta['batch_size']}  (x{meta['grad_accum']} grad accum = {meta['effective_bs']} effective)",
        f"  Learning rate   : {meta['lr']}",
        "",
        "-- Evaluation Results -------------------------------------------",
        f"  Eval loss       : {metrics['eval_loss']}",
        f"  Perplexity      : {metrics['perplexity']}  [{ppl_note}]",
        f"                    (<5 great | 5-15 good | >30 underfit)",
        f"  Token accuracy  : {metrics['token_accuracy']*100:.2f}%  [{tok_note}]",
        f"                    {metrics['correct_tokens']:,} correct / {metrics['total_tokens']:,} answer tokens",
        f"  Wrong tokens    : {metrics['wrong_tokens']:,}",
        f"  Exact match     : {metrics['exact_match']*100:.2f}%  [{ex_note}]",
        f"                    {metrics['exact_matches']} / {metrics['total_samples']} samples",
        f"  Fuzzy match     : {metrics['fuzzy_match']*100:.2f}%",
        f"                    {metrics['fuzzy_matches']} / {metrics['total_samples']} samples",
        f"                    (all reference words found somewhere in prediction)",
        "",
        "-- What These Numbers Mean --------------------------------------",
        "  Token accuracy  : Of every answer token the model was asked to",
        "                    predict, what % did it get right. Measured on",
        "                    answer tokens ONLY (not question/prompt tokens).",
        "  Exact match     : Normalised prediction == normalised reference.",
        "                    Low is normal for generation models -- use Fuzzy.",
        "  Fuzzy match     : All words in the reference appear in prediction.",
        "                    This is the most useful metric for QA.",
        "  Perplexity      : How confident the model is on held-out text.",
        "                    Your 1.4 perplexity = model has learnt the corpus.",
        "",
        "-- Per-Step Train Loss ------------------------------------------",
    ]
    for row in metrics["epoch_losses"]:
        lines.append(f"  epoch {row['epoch']:<6}  step {str(row['step']):<6}  loss {row['train_loss']}")

    lines += ["", "-- Per-Epoch Eval Loss ------------------------------------------"]
    for row in metrics["eval_losses"]:
        lines.append(f"  epoch {row['epoch']:<6}  eval_loss {row['eval_loss']}")

    if metrics.get("sample_log"):
        lines += ["", "-- Sample Predictions (first 10 eval examples) ------------------"]
        for s in metrics["sample_log"]:
            exact_str = "EXACT" if s["exact"] else ("FUZZY" if s["fuzzy"] else "MISS")
            lines += [
                f"  Sample #{s['sample']}  [{exact_str}]  tokens: {s['tok_acc']}",
                f"  REF  : {s['reference']}",
                f"  PRED : {s['predicted']}",
                "",
            ]

    lines += [
        "=" * 66,
    ]

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[METRICS] TXT report saved: {path}", flush=True)


def save_metrics_xlsx(path: Path, metrics: dict, meta: dict) -> None:
    """Save a formatted Excel workbook with 3 sheets: Summary, Loss Curve, Raw."""
    if not HAS_OPENPYXL:
        print("[METRICS] Skipping Excel -- openpyxl not installed.", flush=True)
        return

    wb = openpyxl.Workbook()

    # -- Styles ----------------------------------------------------------------
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill    = PatternFill("solid", fgColor="2E4057")
    sub_fill    = PatternFill("solid", fgColor="4A7C9E")
    good_fill   = PatternFill("solid", fgColor="C8F7C5")
    warn_fill   = PatternFill("solid", fgColor="FFF3CD")
    bad_fill    = PatternFill("solid", fgColor="FADADD")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, row, col, value, fill=None, font=None, number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = center
        c.border    = border
        if fill:   c.fill   = fill
        if font:   c.font   = font
        if number_format: c.number_format = number_format
        return c

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1: Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 22
    ws1.row_dimensions[1].height = 28

    hcell(ws1, 1, 1, "Metric",         hdr_fill, hdr_font)
    hcell(ws1, 1, 2, "Value",          hdr_fill, hdr_font)
    hcell(ws1, 1, 3, "Notes",          hdr_fill, hdr_font)

    summary_rows = [
        # section header
        ("-- Model Config", None, None),
        ("Base Model",       meta["base_model"],      ""),
        ("Device",           meta["device"],           ""),
        ("Precision",        meta["precision"],        "bf16 = best for RTX 30xx"),
        ("Train Time",       f"{meta['train_time_min']} min", ""),
        # section header
        ("-- Dataset", None, None),
        ("QA Pairs Total",   meta["qa_pairs"],         ""),
        ("Training Rows",    meta["train_rows"],        ""),
        ("Eval Rows",        meta["eval_rows"],         ""),
        ("Epochs",           meta["epochs"],            ""),
        ("Batch Size",       meta["batch_size"],        f"x{meta['grad_accum']} = {meta['effective_bs']} effective"),
        ("Learning Rate",    meta["lr"],                "cosine schedule"),
        # section header
        ("-- Eval Metrics", None, None),
        ("Eval Loss",        metrics["eval_loss"],      "lower is better"),
        ("Perplexity",       metrics["perplexity"],     "<5 great | 5-15 good | >30 underfit"),
        ("Token Accuracy",   f"{metrics['token_accuracy']*100:.2f}%", f"{metrics['correct_tokens']:,} / {metrics['total_tokens']:,} tokens"),
        ("Wrong Tokens",     f"{metrics['wrong_tokens']:,}",           ""),
        ("Exact Match",      f"{metrics['exact_match']*100:.2f}%",    f"{metrics['exact_matches']} / {metrics['total_samples']} samples"),
        ("Fuzzy Match",      f"{metrics['fuzzy_match']*100:.2f}%",    f"{metrics['fuzzy_matches']} / {metrics['total_samples']} (all ref words in pred)"),
    ]

    for r, (label, value, note) in enumerate(summary_rows, start=2):
        if value is None:   # section header row
            c = ws1.cell(row=r, column=1, value=label)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = sub_fill
            c.alignment = center
            c.border = border
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws1.row_dimensions[r].height = 20
            continue

        ws1.cell(row=r, column=1, value=label).border  = border
        ws1.cell(row=r, column=2, value=value).border  = border
        ws1.cell(row=r, column=3, value=note).border   = border
        ws1.cell(row=r, column=1).alignment = Alignment(vertical="center")
        ws1.cell(row=r, column=2).alignment = center
        ws1.cell(row=r, column=3).alignment = Alignment(vertical="center")

        # colour-code eval metrics
        if label == "Perplexity":
            fill = good_fill if metrics["perplexity"] < 5 else warn_fill if metrics["perplexity"] < 15 else bad_fill
            ws1.cell(row=r, column=2).fill = fill
        if label == "Token Accuracy":
            pct = metrics["token_accuracy"]
            fill = good_fill if pct > 0.85 else warn_fill if pct > 0.60 else bad_fill
            ws1.cell(row=r, column=2).fill = fill
        if label == "Exact Match":
            pct = metrics["exact_match"]
            fill = good_fill if pct > 0.50 else warn_fill if pct > 0.20 else bad_fill
            ws1.cell(row=r, column=2).fill = fill
        if label == "Fuzzy Match":
            pct = metrics["fuzzy_match"]
            fill = good_fill if pct > 0.60 else warn_fill if pct > 0.30 else bad_fill
            ws1.cell(row=r, column=2).fill = fill

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2: Loss Curve
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Loss Curve")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 18

    hcell(ws2, 1, 1, "Epoch",      hdr_fill, hdr_font)
    hcell(ws2, 1, 2, "Train Loss", hdr_fill, hdr_font)
    hcell(ws2, 1, 3, "Step",       hdr_fill, hdr_font)
    hcell(ws2, 1, 4, "Eval Loss",  hdr_fill, hdr_font)

    # merge train and eval rows by epoch
    eval_by_epoch = {r["epoch"]: r["eval_loss"] for r in metrics["eval_losses"]}
    for r, row in enumerate(metrics["epoch_losses"], start=2):
        hcell(ws2, r, 1, row["epoch"])
        hcell(ws2, r, 2, row["train_loss"])
        hcell(ws2, r, 3, row["step"])
        ev = eval_by_epoch.get(row["epoch"], "")
        hcell(ws2, r, 4, ev)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3: Raw Numbers
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Raw Numbers")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 24

    raw_rows = [
        ("eval_loss",       metrics["eval_loss"]),
        ("perplexity",      metrics["perplexity"]),
        ("token_accuracy",  metrics["token_accuracy"]),
        ("exact_match",     metrics["exact_match"]),
        ("total_tokens",    metrics["total_tokens"]),
        ("correct_tokens",  metrics["correct_tokens"]),
        ("wrong_tokens",    metrics["wrong_tokens"]),
        ("total_samples",   metrics["total_samples"]),
        ("exact_matches",   metrics["exact_matches"]),
        ("train_time_min",  meta["train_time_min"]),
        ("epochs",          meta["epochs"]),
        ("batch_size",      meta["batch_size"]),
        ("grad_accum",      meta["grad_accum"]),
        ("effective_bs",    meta["effective_bs"]),
        ("lr",              meta["lr"]),
        ("precision",       meta["precision"]),
        ("base_model",      meta["base_model"]),
        ("device",          meta["device"]),
        ("qa_pairs",        meta["qa_pairs"]),
        ("train_rows",      meta["train_rows"]),
        ("eval_rows",       meta["eval_rows"]),
    ]
    hcell(ws3, 1, 1, "Key",   hdr_fill, hdr_font)
    hcell(ws3, 1, 2, "Value", hdr_fill, hdr_font)
    for r, (k, v) in enumerate(raw_rows, start=2):
        ws3.cell(row=r, column=1, value=k).border   = border
        ws3.cell(row=r, column=2, value=v).border   = border
        ws3.cell(row=r, column=1).alignment = Alignment(vertical="center")
        ws3.cell(row=r, column=2).alignment = center

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 4: Sample Predictions
    # ══════════════════════════════════════════════════════════════════════════
    if metrics.get("sample_log"):
        ws4 = wb.create_sheet("Sample Predictions")
        ws4.column_dimensions["A"].width = 8
        ws4.column_dimensions["B"].width = 60
        ws4.column_dimensions["C"].width = 60
        ws4.column_dimensions["D"].width = 12
        ws4.column_dimensions["E"].width = 12
        ws4.column_dimensions["F"].width = 12

        headers = ["#", "Reference Answer", "Predicted Answer", "Exact", "Fuzzy", "Tok Acc"]
        for col, h in enumerate(headers, 1):
            hcell(ws4, 1, col, h, hdr_fill, hdr_font)

        for r, s in enumerate(metrics["sample_log"], start=2):
            result_fill = good_fill if s["exact"] else (warn_fill if s["fuzzy"] else bad_fill)
            ws4.cell(row=r, column=1, value=s["sample"]).border = border
            c_ref  = ws4.cell(row=r, column=2, value=s["reference"])
            c_pred = ws4.cell(row=r, column=3, value=s["predicted"])
            c_ref.border  = border
            c_pred.border = border
            c_ref.alignment  = Alignment(wrap_text=True, vertical="top")
            c_pred.alignment = Alignment(wrap_text=True, vertical="top")
            c_pred.fill = result_fill
            hcell(ws4, r, 4, "YES" if s["exact"] else "NO",  good_fill if s["exact"] else bad_fill)
            hcell(ws4, r, 5, "YES" if s["fuzzy"] else "NO",  good_fill if s["fuzzy"] else warn_fill)
            hcell(ws4, r, 6, s["tok_acc"])
            ws4.row_dimensions[r].height = 45

    wb.save(path)
    print(f"[METRICS] Excel report saved: {path}", flush=True)


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main() -> None:
    # -- Validate paths --------------------------------------------------------
    if not CORPUS_PATH.exists():
        sys.exit(
            f"[ERROR] Corpus not found: {CORPUS_PATH}\n"
            "Edit CORPUS_PATH at the top of train.py"
        )

    # -- Device + precision ----------------------------------------------------
    device = detect_device()
    use_fp16, use_bf16, torch_dtype = pick_precision(device)

    # -- Dataset ---------------------------------------------------------------
    rows = load_or_build_dataset(DATASET_PATH, CORPUS_PATH)
    if not rows:
        sys.exit("[ERROR] No training rows generated. Check your corpus file.")

    hf_dataset = Dataset.from_list([{"text": format_example(r)} for r in rows])
    test_size  = min(0.1, max(1 / len(rows), 0.05))
    split      = hf_dataset.train_test_split(test_size=test_size, seed=42)

    # -- Tokenizer -------------------------------------------------------------
    print(f"[MODEL] Loading tokenizer: {BASE_MODEL}", flush=True)
    tokenizer = AutoTokenizer.from_pretrained(BASE_MODEL, use_fast=True)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token

    # -- Model -----------------------------------------------------------------
    print(f"[MODEL] Loading base model: {BASE_MODEL} (dtype={torch_dtype})", flush=True)
    model = AutoModelForCausalLM.from_pretrained(
        BASE_MODEL,
        dtype=torch_dtype,
        device_map="auto" if device == "cuda" else None,
    )
    model.gradient_checkpointing_enable()
    model.config.use_cache = False   # required with gradient checkpointing

    # -- Tokenise --------------------------------------------------------------
    def preprocess(batch: dict) -> dict:
        enc = tokenizer(
            batch["text"],
            truncation=True,
            max_length=MAX_LENGTH,
            padding="max_length",
        )
        enc["labels"] = enc["input_ids"].copy()
        return enc

    num_workers  = min(4, os.cpu_count() or 1) if device != "mps" else 0
    tok_train    = split["train"].map(preprocess, batched=True, num_proc=num_workers,
                                      remove_columns=split["train"].column_names)
    tok_eval     = split["test"].map(preprocess,  batched=True, num_proc=num_workers,
                                      remove_columns=split["test"].column_names)
    tok_train.set_format("torch")
    tok_eval.set_format("torch")

    # -- TrainingArguments -----------------------------------------------------
    # Use a TEMP dir for checkpoints so OUTPUT_DIR stays clean.
    # After training we save the final model directly to OUTPUT_DIR.
    tmp_ckpt_dir = OUTPUT_DIR.parent / (OUTPUT_DIR.name + "_checkpoints")
    tmp_ckpt_dir.mkdir(parents=True, exist_ok=True)

    supported = inspect.signature(TrainingArguments.__init__).parameters

    arg_map: dict = {
        "output_dir":                   str(tmp_ckpt_dir),
        "num_train_epochs":             EPOCHS,
        "per_device_train_batch_size":  BATCH_SIZE,
        "per_device_eval_batch_size":   BATCH_SIZE,
        "gradient_accumulation_steps":  GRAD_ACCUM,
        "learning_rate":                LR,
        "warmup_ratio":                 WARMUP_RATIO,
        "weight_decay":                 0.01,
        "max_grad_norm":                1.0,
        "fp16":                         use_fp16,
        "bf16":                         use_bf16,
        "dataloader_num_workers":       num_workers,
        "logging_steps":                10,
        # ---- KEY CHANGE: save nothing during training ----
        "save_strategy":                "no",
        "report_to":                    "none",
    }

    # Optional args guarded for HF version compatibility
    optional: dict = {
        "lr_scheduler_type":          "cosine",
        "dataloader_pin_memory":      device == "cuda",
        "ddp_find_unused_parameters": False,
        "group_by_length":            True,
        "optim":                      "adamw_torch_fused" if device == "cuda" else "adamw_torch",
    }
    # tf32 only works on CUDA Ampere+ (sm_80+)
    if device == "cuda":
        cap = torch.cuda.get_device_capability(0)
        if cap[0] >= 8:
            optional["tf32"] = True
    for k, v in optional.items():
        if k in supported:
            arg_map[k] = v

    # warmup_ratio was deprecated in newer HF; fall back to warmup_steps
    if "warmup_ratio" not in supported and "warmup_steps" in supported:
        arg_map.pop("warmup_ratio", None)
        total_steps = (len(tok_train) // (BATCH_SIZE * GRAD_ACCUM)) * EPOCHS
        arg_map["warmup_steps"] = max(1, int(total_steps * WARMUP_RATIO))

    if "evaluation_strategy" in supported:
        arg_map["evaluation_strategy"] = "epoch"
    elif "eval_strategy" in supported:
        arg_map["eval_strategy"] = "epoch"

    if "overwrite_output_dir" in supported:
        arg_map["overwrite_output_dir"] = True

    training_args = TrainingArguments(**arg_map)

    # -- Trainer ---------------------------------------------------------------
    trainer = Trainer(
        model=model,
        args=training_args,
        train_dataset=tok_train,
        eval_dataset=tok_eval,
        data_collator=DataCollatorForLanguageModeling(tokenizer=tokenizer, mlm=False),
    )

    # -- Summary banner --------------------------------------------------------
    eff_bs = BATCH_SIZE * GRAD_ACCUM
    vram_info = (
        f"{torch.cuda.get_device_properties(0).total_memory/1e9:.1f} GB"
        if device == "cuda" else "N/A"
    )
    print(f"\n{'='*60}")
    print(f"  Base model           : {BASE_MODEL}")
    print(f"  GPU / VRAM           : {torch.cuda.get_device_name(0) if device=='cuda' else device}  /  {vram_info}")
    print(f"  Training samples     : {len(tok_train)}")
    print(f"  Eval samples         : {len(tok_eval)}")
    print(f"  Epochs               : {EPOCHS}")
    print(f"  Effective batch size : {eff_bs}  ({BATCH_SIZE} x {GRAD_ACCUM} grad_accum)")
    print(f"  Learning rate        : {LR}  (cosine + {WARMUP_RATIO:.0%} warmup)")
    print(f"  Mixed precision      : {'bf16 (Ampere)' if use_bf16 else 'fp16' if use_fp16 else 'fp32 (CPU)'}")
    print(f"  Gradient checkpt.    : enabled  (saves ~30% VRAM)")
    print(f"  Saving to            : {OUTPUT_DIR}")
    print(f"  OOM tip              : if you get CUDA OOM, set BATCH_SIZE=1 in CONFIG")
    print(f"{'='*60}\n")

    # -- Train -----------------------------------------------------------------
    t0 = time.time()
    trainer.train()
    elapsed = time.time() - t0
    print(f"\n[TRAIN] Finished in {elapsed/60:.1f} min", flush=True)

    # -- Save ONE clean final model --------------------------------------------
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"[SAVE] Writing final model to {OUTPUT_DIR} ...", flush=True)
    # Re-enable cache for inference before saving
    model.config.use_cache = True
    trainer.save_model(str(OUTPUT_DIR))
    tokenizer.save_pretrained(str(OUTPUT_DIR))

    # Remove the temp checkpoint dir entirely
    if tmp_ckpt_dir.exists():
        shutil.rmtree(tmp_ckpt_dir)
        print(f"[SAVE] Removed temp checkpoint dir.", flush=True)

    # -- Metadata dict (shared by metrics + JSON) -----------------------------
    meta = {
        "base_model":     BASE_MODEL,
        "output_dir":     str(OUTPUT_DIR),
        "corpus":         str(CORPUS_PATH),
        "dataset":        str(DATASET_PATH),
        "qa_pairs":       len(rows),
        "train_rows":     len(tok_train),
        "eval_rows":      len(tok_eval),
        "epochs":         EPOCHS,
        "batch_size":     BATCH_SIZE,
        "grad_accum":     GRAD_ACCUM,
        "effective_bs":   eff_bs,
        "lr":             LR,
        "max_length":     MAX_LENGTH,
        "precision":      "bf16" if use_bf16 else "fp16" if use_fp16 else "fp32",
        "device":         device,
        "train_time_min": round(elapsed / 60, 1),
    }
    (OUTPUT_DIR / "training_metadata.json").write_text(
        json.dumps(meta, indent=2), encoding="utf-8"
    )

    # -- Evaluate + save metrics -----------------------------------------------
    metrics = compute_eval_metrics(trainer, tok_eval, tokenizer, device)

    # Print summary to console
    print(f"\n[METRICS] -- Results ----------------------------------------------")
    print(f"[METRICS]  Perplexity    : {metrics['perplexity']}  (<5 great | 5-15 good | >30 underfit)")
    print(f"[METRICS]  Token Acc.    : {metrics['token_accuracy']*100:.2f}%  "
          f"({metrics['correct_tokens']:,} correct / {metrics['total_tokens']:,} answer tokens)")
    print(f"[METRICS]  Wrong tokens  : {metrics['wrong_tokens']:,}")
    print(f"[METRICS]  Exact Match   : {metrics['exact_match']*100:.2f}%  ({metrics['exact_matches']} / {metrics['total_samples']})")
    print(f"[METRICS]  Fuzzy Match   : {metrics['fuzzy_match']*100:.2f}%  ({metrics['fuzzy_matches']} / {metrics['total_samples']}  -- all ref words found in pred)")
    print(f"[METRICS] ----------------------------------------------------------\n")

    # Save TXT report (always)
    txt_path = OUTPUT_DIR / "training_metrics.txt"
    save_metrics_txt(txt_path, metrics, meta)

    # Save Excel report (if openpyxl available)
    xlsx_path = OUTPUT_DIR / "training_metrics.xlsx"
    save_metrics_xlsx(xlsx_path, metrics, meta)

    # Save raw metrics JSON too
    (OUTPUT_DIR / "training_metrics.json").write_text(
        json.dumps({**meta, **metrics}, indent=2, default=str), encoding="utf-8"
    )
    print(f"[METRICS] JSON saved:  {OUTPUT_DIR / 'training_metrics.json'}", flush=True)

    print(f"\n{'='*60}")
    print(f"  DONE.  Model saved to : {OUTPUT_DIR}")
    print(f"  Metrics TXT           : {txt_path.name}")
    if HAS_OPENPYXL:
        print(f"  Metrics Excel         : {xlsx_path.name}")
    print(f"  Run: python ask.py \"what is the fee structure\"")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()